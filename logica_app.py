# -*- coding: utf-8 -*-
"""
Módulo de lógica de negócio
Contém funções para processamento de dados e PDF
Versão SQLite com otimizações + Histórico de Faltas
"""

import sqlite3
import pandas as pd
import fitz  # PyMuPDF
from datetime import datetime
import re
import os

# Caminhos dos bancos de dados
DB_ALUNOS = 'data/alunos.db'
DB_HORARIOS = 'data/horarios.db'
DB_FALTAS = 'data/faltas_consolidadas.db'
DB_HISTORICO = 'data/historico_faltas.db'  # NOVO!

# ==================== INICIALIZAÇÃO DOS BANCOS ====================

def inicializar_bancos():
    """Cria os bancos de dados e tabelas se não existirem"""
    
    os.makedirs('data', exist_ok=True)
    
    # Banco de Alunos
    conn = sqlite3.connect(DB_ALUNOS)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alunos (
            matricula TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            turma TEXT NOT NULL
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_nome ON alunos(nome)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_turma ON alunos(turma)')
    conn.commit()
    conn.close()
    
    # Banco de Horários
    conn = sqlite3.connect(DB_HORARIOS)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS horarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            turma TEXT NOT NULL,
            dia_semana TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            disciplina TEXT NOT NULL
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_turma_dia ON horarios(turma, dia_semana)')
    conn.commit()
    conn.close()
    
    # Banco de Faltas Consolidadas
    conn = sqlite3.connect(DB_FALTAS)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS faltas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula_aluno TEXT NOT NULL,
            disciplina TEXT NOT NULL,
            total_faltas INTEGER DEFAULT 0,
            UNIQUE(matricula_aluno, disciplina)
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_matricula ON faltas(matricula_aluno)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_disciplina ON faltas(disciplina)')
    conn.commit()
    conn.close()
    
    # NOVO: Banco de Histórico de Faltas
    conn = sqlite3.connect(DB_HISTORICO)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS historico_faltas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula_aluno TEXT NOT NULL,
            nome_aluno TEXT NOT NULL,
            turma TEXT NOT NULL,
            disciplina TEXT NOT NULL,
            data TEXT NOT NULL,
            dia_semana TEXT NOT NULL,
            UNIQUE(matricula_aluno, disciplina, data)
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_hist_matricula ON historico_faltas(matricula_aluno)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_hist_data ON historico_faltas(data)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_hist_turma ON historico_faltas(turma)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_hist_disciplina ON historico_faltas(disciplina)')
    conn.commit()
    conn.close()


# ==================== FUNÇÕES DE LEITURA DE PDF ====================

def extrair_matriculas_presentes(caminho_pdf):
    """Extrai matrículas de alunos presentes do PDF de frequência"""
    try:
        doc = fitz.open(caminho_pdf)
        texto_completo = ""
        
        for pagina in doc:
            texto_completo += pagina.get_text()
        
        doc.close()
        
        linhas = texto_completo.split('\n')
        matriculas = []
        
        for i, linha in enumerate(linhas):
            if 'Crachá:' in linha and 'Nome:' in linha:
                if i + 1 < len(linhas):
                    proxima_linha = linhas[i + 1].strip()
                    match = re.match(r'^(\d{10})\s+', proxima_linha)
                    if match:
                        matriculas.append(match.group(1))
        
        return list(set(matriculas))
        
    except Exception as e:
        raise Exception(f"Erro ao ler PDF de presentes: {str(e)}")


def extrair_matriculas_ausentes(caminho_pdf):
    """Extrai matrículas de alunos ausentes do PDF"""
    try:
        doc = fitz.open(caminho_pdf)
        texto_completo = ""
        
        for pagina in doc:
            texto_completo += pagina.get_text()
        
        doc.close()
        
        matriculas = re.findall(r'^(\d{10})\s+[A-ZÀÁÂÃÉÊÍÓÔÕÚÇ\s]+ALUNO', texto_completo, re.MULTILINE)
        
        return list(set(matriculas))
        
    except Exception as e:
        raise Exception(f"Erro ao ler PDF de ausentes: {str(e)}")


# ==================== FUNÇÕES CRUD - ALUNOS ====================

def carregar_alunos(limite=None, offset=0, termo_busca=None, filtro_turma=None):
    """Carrega os dados de alunos com paginação e busca"""
    try:
        conn = sqlite3.connect(DB_ALUNOS)
        
        conditions = []
        params = []
        
        if termo_busca:
            conditions.append("(matricula LIKE ? OR nome LIKE ?)")
            params.extend([f'%{termo_busca}%', f'%{termo_busca}%'])
        
        if filtro_turma and filtro_turma != "Todas":
            conditions.append("turma = ?")
            params.append(filtro_turma)
        
        query = 'SELECT matricula, nome, turma FROM alunos'
        
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        
        query += ' ORDER BY nome'
        
        if limite:
            query += ' LIMIT ? OFFSET ?'
            params.extend([limite, offset])
        
        if params:
            df = pd.read_sql_query(query, conn, params=params)
        else:
            df = pd.read_sql_query(query, conn)
        
        conn.close()
        return df
        
    except Exception as e:
        raise Exception(f"Erro ao carregar alunos: {str(e)}")


def contar_alunos(termo_busca=None, filtro_turma=None):
    """Retorna o total de alunos (para paginação)"""
    try:
        conn = sqlite3.connect(DB_ALUNOS)
        cursor = conn.cursor()
        
        conditions = []
        params = []
        
        if termo_busca:
            conditions.append("(matricula LIKE ? OR nome LIKE ?)")
            params.extend([f'%{termo_busca}%', f'%{termo_busca}%'])
        
        if filtro_turma and filtro_turma != "Todas":
            conditions.append("turma = ?")
            params.append(filtro_turma)
        
        query = 'SELECT COUNT(*) FROM alunos'
        
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute('SELECT COUNT(*) FROM alunos')
        
        total = cursor.fetchone()[0]
        conn.close()
        return total
        
    except Exception as e:
        return 0


def adicionar_aluno(dados_aluno):
    """Adiciona um novo aluno ao banco de dados"""
    try:
        conn = sqlite3.connect(DB_ALUNOS)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO alunos (matricula, nome, turma)
            VALUES (?, ?, ?)
        ''', (dados_aluno['matricula'], dados_aluno['nome'], dados_aluno['turma']))
        
        conn.commit()
        conn.close()
        
    except sqlite3.IntegrityError:
        raise Exception("Matrícula já cadastrada!")
    except Exception as e:
        raise Exception(f"Erro ao adicionar aluno: {str(e)}")


def atualizar_aluno(matricula, novos_dados):
    """Atualiza os dados de um aluno existente"""
    try:
        conn = sqlite3.connect(DB_ALUNOS)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE alunos 
            SET matricula = ?, nome = ?, turma = ?
            WHERE matricula = ?
        ''', (novos_dados['matricula'], novos_dados['nome'], 
              novos_dados['turma'], matricula))
        
        if cursor.rowcount == 0:
            raise Exception("Aluno não encontrado!")
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao atualizar aluno: {str(e)}")


def excluir_aluno(matricula):
    """Remove um aluno do banco de dados"""
    try:
        conn = sqlite3.connect(DB_ALUNOS)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM alunos WHERE matricula = ?', (matricula,))
        
        if cursor.rowcount == 0:
            raise Exception("Aluno não encontrado!")
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao excluir aluno: {str(e)}")


# ==================== FUNÇÕES CRUD - HORÁRIOS ====================

def carregar_horarios(limite=None, offset=0, turma_filtro=None, dia_semana_filtro=None):
    """Carrega os dados de horários com paginação"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        
        conditions = []
        params = []
        
        if turma_filtro and turma_filtro != "Todas":
            conditions.append("turma = ?")
            params.append(turma_filtro)
        
        if dia_semana_filtro and dia_semana_filtro != "Todos":
            conditions.append("dia_semana = ?")
            params.append(dia_semana_filtro)
        
        query = '''
            SELECT id, turma, dia_semana, hora_inicio, hora_fim, disciplina 
            FROM horarios
        '''
        
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        
        query += ' ORDER BY turma, dia_semana, hora_inicio'
        
        if limite:
            query += ' LIMIT ? OFFSET ?'
            params.extend([limite, offset])
        
        if params:
            df = pd.read_sql_query(query, conn, params=params)
        else:
            df = pd.read_sql_query(query, conn)
        
        conn.close()
        return df
        
    except Exception as e:
        raise Exception(f"Erro ao carregar horários: {str(e)}")


def contar_horarios(turma_filtro=None, dia_semana_filtro=None):
    """Retorna o total de horários (para paginação)"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        
        conditions = []
        params = []
        
        if turma_filtro and turma_filtro != "Todas":
            conditions.append("turma = ?")
            params.append(turma_filtro)
        
        if dia_semana_filtro and dia_semana_filtro != "Todos":
            conditions.append("dia_semana = ?")
            params.append(dia_semana_filtro)
        
        query = 'SELECT COUNT(*) FROM horarios'
        
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute('SELECT COUNT(*) FROM horarios')
        
        total = cursor.fetchone()[0]
        conn.close()
        return total
        
    except Exception as e:
        return 0

def obter_dias_semana_distintos():
    """Retorna lista de dias da semana padronizados"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT UPPER(TRIM(dia_semana)) FROM horarios')
        dias_raw = [row[0] for row in cursor.fetchall()]
        conn.close()

        dias_padrao = [
            "SEGUNDA-FEIRA",
            "TERÇA-FEIRA",
            "QUARTA-FEIRA",
            "QUINTA-FEIRA",
            "SEXTA-FEIRA",
            "SÁBADO",
            "DOMINGO"
        ]

        return [d for d in dias_padrao if d in dias_raw] or dias_padrao
    except:
        return ["SEGUNDA-FEIRA", "TERÇA-FEIRA", "QUARTA-FEIRA", "QUINTA-FEIRA", "SEXTA-FEIRA", "SÁBADO", "DOMINGO"]


def adicionar_horario(dados_horario):
    """Adiciona um novo horário ao banco de dados"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO horarios (turma, dia_semana, hora_inicio, hora_fim, disciplina)
            VALUES (?, ?, ?, ?, ?)
        ''', (dados_horario['turma'], dados_horario['dia_semana'],
              dados_horario['hora_inicio'], dados_horario['hora_fim'],
              dados_horario['disciplina']))
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao adicionar horário: {str(e)}")


def atualizar_horario(id_horario, novos_dados):
    """Atualiza os dados de um horário existente"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE horarios 
            SET turma = ?, dia_semana = ?, hora_inicio = ?, hora_fim = ?, disciplina = ?
            WHERE id = ?
        ''', (novos_dados['turma'], novos_dados['dia_semana'],
              novos_dados['hora_inicio'], novos_dados['hora_fim'],
              novos_dados['disciplina'], id_horario))
        
        if cursor.rowcount == 0:
            raise Exception("Horário não encontrado!")
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao atualizar horário: {str(e)}")


def excluir_horario(id_horario):
    """Remove um horário do banco de dados"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM horarios WHERE id = ?', (id_horario,))
        
        if cursor.rowcount == 0:
            raise Exception("Horário não encontrado!")
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao excluir horário: {str(e)}")


def obter_turmas_distintas():
    """Retorna lista de turmas únicas"""
    try:
        conn = sqlite3.connect(DB_ALUNOS)
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT turma FROM alunos ORDER BY turma')
        turmas = [row[0] for row in cursor.fetchall()]
        conn.close()
        return turmas
    except:
        return []


# ==================== FUNÇÕES DE FALTAS ====================

def converter_dia_semana(data):
    """Converte data para o formato de dia da semana usado no banco"""
    dias = {
        0: 'SEGUNDA-FEIRA',
        1: 'TERÇA-FEIRA',
        2: 'QUARTA-FEIRA',
        3: 'QUINTA-FEIRA',
        4: 'SEXTA-FEIRA',
        5: 'SÁBADO',
        6: 'DOMINGO'
    }
    return dias[data.weekday()]


def processar_faltas_do_dia(data_str, lista_presentes, lista_ausentes):
    """Processa as faltas de um dia específico - ATUALIZADO para salvar no histórico"""
    try:
        # Converter data
        try:
            data = datetime.strptime(data_str, '%d/%m/%Y')
        except:
            data = datetime.strptime(data_str, '%Y-%m-%d')
        
        data_formatada = data.strftime('%Y-%m-%d')  # Para o banco
        dia_semana = converter_dia_semana(data)
        
        # Carregar dados
        conn_alunos = sqlite3.connect(DB_ALUNOS)
        df_alunos = pd.read_sql_query('SELECT matricula, nome, turma FROM alunos', conn_alunos)
        conn_alunos.close()
        
        # Normalizar matrículas dos PDFs
        lista_presentes_norm = [str(m).lstrip('0') for m in lista_presentes if m]
        lista_ausentes_norm = [str(m).lstrip('0') for m in lista_ausentes if m]
        
        # Criar dicionário: matrícula_normalizada -> matrícula_original
        mapa_matriculas = {}
        for _, aluno in df_alunos.iterrows():
            mat_original = str(aluno['matricula'])
            mat_norm = mat_original.lstrip('0')
            mapa_matriculas[mat_norm] = mat_original
        
        conn_horarios = sqlite3.connect(DB_HORARIOS)
        df_horarios = pd.read_sql_query(
            'SELECT turma, disciplina FROM horarios WHERE dia_semana = ?',
            conn_horarios, params=[dia_semana]
        )
        conn_horarios.close()
        
        # Determinar ausentes finais
        ausentes_finais_norm = set(lista_ausentes_norm)
        
        for mat_norm in mapa_matriculas.keys():
            if mat_norm not in lista_presentes_norm:
                ausentes_finais_norm.add(mat_norm)
        
        faltas_dia = []
        
        # Conectar aos bancos de faltas
        conn_faltas = sqlite3.connect(DB_FALTAS)
        cursor_faltas = conn_faltas.cursor()
        
        conn_historico = sqlite3.connect(DB_HISTORICO)
        cursor_historico = conn_historico.cursor()
        
        # Processar cada ausente
        for matricula_norm in ausentes_finais_norm:
            if matricula_norm not in mapa_matriculas:
                continue
            
            matricula_original = mapa_matriculas[matricula_norm]
            aluno = df_alunos[df_alunos['matricula'] == matricula_original]
            
            if len(aluno) == 0:
                continue
            
            turma = aluno.iloc[0]['turma']
            nome = aluno.iloc[0]['nome']
            
            # Disciplinas da turma nesse dia
            disciplinas_dia = df_horarios[df_horarios['turma'] == turma]
            
            if len(disciplinas_dia) == 0:
                continue
            
            # Registrar falta em cada disciplina
            for _, horario in disciplinas_dia.iterrows():
                disciplina = horario['disciplina']
                
                faltas_dia.append({
                    'matricula': matricula_original,
                    'nome': nome,
                    'turma': turma,
                    'disciplina': disciplina,
                    'data': data_str,
                    'dia_semana': dia_semana
                })
                
                # 1. Atualizar banco CONSOLIDADO (como antes)
                cursor_faltas.execute('''
                    INSERT INTO faltas (matricula_aluno, disciplina, total_faltas)
                    VALUES (?, ?, 1)
                    ON CONFLICT(matricula_aluno, disciplina) 
                    DO UPDATE SET total_faltas = total_faltas + 1
                ''', (matricula_original, disciplina))
                
                # 2. NOVO: Salvar no HISTÓRICO
                try:
                    cursor_historico.execute('''
                        INSERT INTO historico_faltas 
                        (matricula_aluno, nome_aluno, turma, disciplina, data, dia_semana)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (matricula_original, nome, turma, disciplina, data_formatada, dia_semana))
                except sqlite3.IntegrityError:
                    # Falta já registrada neste dia/disciplina (evita duplicatas)
                    pass
        
        conn_faltas.commit()
        conn_faltas.close()
        
        conn_historico.commit()
        conn_historico.close()
        
        return faltas_dia
        
    except Exception as e:
        raise Exception(f"Erro ao processar faltas: {str(e)}")


def obter_quadro_geral_faltas(limite=None, offset=0, filtro_nome=None, filtro_disciplina=None, filtro_turma=None, filtro_dia_semana=None):
    """Retorna o quadro geral de faltas com paginação e filtros"""
    try:
        conn_faltas = sqlite3.connect(DB_FALTAS)
        cursor = conn_faltas.cursor()
        
        cursor.execute(f"ATTACH DATABASE '{DB_ALUNOS}' AS db_alunos")
        cursor.execute(f"ATTACH DATABASE '{DB_HORARIOS}' AS db_horarios")
        
        query = '''
            SELECT 
                f.matricula_aluno, 
                a.nome,
                a.turma,
                f.disciplina, 
                f.total_faltas
            FROM faltas AS f
            LEFT JOIN db_alunos.alunos AS a ON f.matricula_aluno = a.matricula
        '''
        
        conditions = []
        params = []
        
        if filtro_nome:
            conditions.append("a.nome LIKE ?")
            params.append(f'%{filtro_nome}%')
        
        if filtro_disciplina and filtro_disciplina != "Todas":
            conditions.append("f.disciplina = ?")
            params.append(filtro_disciplina)
        
        if filtro_turma and filtro_turma != "Todas":
            conditions.append("a.turma = ?")
            params.append(filtro_turma)
        
        if filtro_dia_semana and filtro_dia_semana != "Todos":
            conditions.append('''
                EXISTS (
                    SELECT 1 FROM db_horarios.horarios AS h 
                    WHERE h.disciplina = f.disciplina 
                    AND h.turma = a.turma 
                    AND h.dia_semana = ?
                )
            ''')
            params.append(filtro_dia_semana)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY f.total_faltas DESC"
        
        if limite:
            query += ' LIMIT ? OFFSET ?'
            params.extend([limite, offset])

        if params:
            df = pd.read_sql_query(query, conn_faltas, params=params)
        else:
            df = pd.read_sql_query(query, conn_faltas)
        
        df.columns = ['Matrícula', 'Nome', 'Turma', 'Disciplina', 'Total de Faltas']
        df['Nome'] = df['Nome'].fillna('Aluno não encontrado')
        df['Turma'] = df['Turma'].fillna('-')
        
        conn_faltas.close()
        
        return df
        
    except Exception as e:
        raise Exception(f"Erro ao obter quadro geral: {str(e)}")


def contar_faltas(filtro_nome=None, filtro_disciplina=None, filtro_turma=None, filtro_dia_semana=None):
    """Retorna o total de registros de faltas (para paginação) com filtros"""
    try:
        conn = sqlite3.connect(DB_FALTAS)
        cursor = conn.cursor()
        
        cursor.execute(f"ATTACH DATABASE '{DB_ALUNOS}' AS db_alunos")
        cursor.execute(f"ATTACH DATABASE '{DB_HORARIOS}' AS db_horarios")
        
        query = '''
            SELECT COUNT(*) 
            FROM faltas AS f
            LEFT JOIN db_alunos.alunos AS a ON f.matricula_aluno = a.matricula
        '''
        
        conditions = []
        params = []
        
        if filtro_nome:
            conditions.append("a.nome LIKE ?")
            params.append(f'%{filtro_nome}%')
        
        if filtro_disciplina and filtro_disciplina != "Todas":
            conditions.append("f.disciplina = ?")
            params.append(filtro_disciplina)
        
        if filtro_turma and filtro_turma != "Todas":
            conditions.append("a.turma = ?")
            params.append(filtro_turma)
        
        if filtro_dia_semana and filtro_dia_semana != "Todos":
            conditions.append('''
                EXISTS (
                    SELECT 1 FROM db_horarios.horarios AS h 
                    WHERE h.disciplina = f.disciplina 
                    AND h.turma = a.turma 
                    AND h.dia_semana = ?
                )
            ''')
            params.append(filtro_dia_semana)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        
        total = cursor.fetchone()[0]
        conn.close()
        return total
    except:
        return 0

def adicionar_falta(dados_falta):
    """Adiciona uma nova falta manualmente"""
    try:
        conn = sqlite3.connect(DB_FALTAS)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO faltas (matricula_aluno, disciplina, total_faltas)
            VALUES (?, ?, ?)
            ON CONFLICT(matricula_aluno, disciplina) 
            DO UPDATE SET total_faltas = total_faltas + ?
        ''', (dados_falta['matricula'], dados_falta['disciplina'], 
              dados_falta['total_faltas'], dados_falta['total_faltas']))
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao adicionar falta: {str(e)}")


def atualizar_falta(matricula, disciplina, novo_total):
    """Atualiza o total de faltas de um aluno em uma disciplina"""
    try:
        conn = sqlite3.connect(DB_FALTAS)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE faltas 
            SET total_faltas = ?
            WHERE matricula_aluno = ? AND disciplina = ?
        ''', (novo_total, matricula, disciplina))
        
        if cursor.rowcount == 0:
            raise Exception("Registro de falta não encontrado!")
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao atualizar falta: {str(e)}")


def excluir_falta(matricula, disciplina):
    """Remove um registro de falta"""
    try:
        conn = sqlite3.connect(DB_FALTAS)
        cursor = conn.cursor()
        
        cursor.execute('''
            DELETE FROM faltas 
            WHERE matricula_aluno = ? AND disciplina = ?
        ''', (matricula, disciplina))
        
        if cursor.rowcount == 0:
            raise Exception("Registro de falta não encontrado!")
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        raise Exception(f"Erro ao excluir falta: {str(e)}")


def obter_disciplinas_distintas():
    """Retorna lista de disciplinas únicas"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT disciplina FROM horarios ORDER BY disciplina')
        disciplinas = [row[0] for row in cursor.fetchall()]
        conn.close()
        return disciplinas
    except:
        return []
        
def limpar_todas_faltas():
    """Remove todos os registros de faltas do banco (zerar para novo período)"""
    try:
        # Limpar banco consolidado
        conn = sqlite3.connect(DB_FALTAS)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM faltas')
        conn.commit()
        conn.close()
        
        # Limpar histórico
        conn = sqlite3.connect(DB_HISTORICO)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM historico_faltas')
        conn.commit()
        conn.close()
    except Exception as e:
        raise Exception(f"Erro ao limpar faltas: {str(e)}")
        
def obter_horarios_por_turma_grade(turma):
    """Retorna horários organizados em formato de grade para visualização"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        df = pd.read_sql_query(
            'SELECT id, turma, dia_semana, hora_inicio, hora_fim, disciplina FROM horarios WHERE turma = ? ORDER BY dia_semana, hora_inicio',
            conn, params=[turma]
        )
        conn.close()
        return df
    except Exception as e:
        raise Exception(f"Erro ao carregar horários: {str(e)}")

def obter_horarios_todas_turmas_grade():
    """Retorna todos os horários organizados para grade"""
    try:
        conn = sqlite3.connect(DB_HORARIOS)
        df = pd.read_sql_query(
            'SELECT id, turma, dia_semana, hora_inicio, hora_fim, disciplina FROM horarios ORDER BY turma, dia_semana, hora_inicio',
            conn
        )
        conn.close()
        return df
    except Exception as e:
        raise Exception(f"Erro ao carregar horários: {str(e)}")

def obter_horario_por_id(id_horario):
    """Retorna um horário específico pelo ID"""
    try:
        id_horario = int(str(id_horario).strip())
        conn = sqlite3.connect(DB_HORARIOS)
        cursor = conn.cursor()
        cursor.execute('SELECT id, turma, dia_semana, hora_inicio, hora_fim, disciplina FROM horarios WHERE id = ?', (id_horario,))
        horario = cursor.fetchone()
        conn.close()
        
        if horario is None:
            raise Exception(f"Horário com ID {id_horario} não encontrado")
        
        return horario
    except Exception as e:
<<<<<<< HEAD
        raise Exception(f"Erro ao buscar horário: {str(e)}")


# ==================== NOVAS FUNÇÕES: EXPORTAÇÃO EXCEL ====================

def exportar_faltas_para_excel(mes, ano, caminho_arquivo):
    """
    Exporta faltas do mês/ano especificado para Excel no formato da planilha
    
    Args:
        mes: Número do mês (1-12)
        ano: Ano (ex: 2025)
        caminho_arquivo: Caminho completo onde salvar o arquivo .xlsx
    """
    try:
        import calendar
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Obter dados do histórico do mês
        data_inicio = f"{ano}-{mes:02d}-01"
        dias_no_mes = calendar.monthrange(ano, mes)[1]
        data_fim = f"{ano}-{mes:02d}-{dias_no_mes}"
        
        conn = sqlite3.connect(DB_HISTORICO)
        
        # Buscar todas as faltas do mês
        query = '''
            SELECT matricula_aluno, nome_aluno, turma, disciplina, data
            FROM historico_faltas
            WHERE data BETWEEN ? AND ?
            ORDER BY turma, nome_aluno, data
        '''
        
        df_historico = pd.read_sql_query(query, conn, params=[data_inicio, data_fim])
        conn.close()
        
        if df_historico.empty:
            raise Exception(f"Nenhuma falta registrada para {mes:02d}/{ano}")
        
        # Buscar todas as disciplinas do sistema
        disciplinas_sistema = obter_disciplinas_distintas()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        
        # Nome do mês
        meses = ['', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
                 'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        ws.title = f"{meses[mes]} {ano}"
        
        # Estilos
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        total_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # LINHA 1: Título do mês
        ws.merge_cells('A1:C1')
        cell_titulo = ws['A1']
        cell_titulo.value = meses[mes]
        cell_titulo.fill = header_fill
        cell_titulo.font = Font(bold=True, color="FFFFFF", size=14)
        cell_titulo.alignment = center_align
        
        # Cabeçalhos das colunas de dias
        col_offset = 4  # Começa na coluna D (depois de TURMA, NOME, vazio)
        for dia in range(1, dias_no_mes + 1):
            cell = ws.cell(row=1, column=col_offset + dia - 1)
            cell.value = dia
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 3
        
        # Cabeçalhos das disciplinas
        col_disciplinas = col_offset + dias_no_mes
        for idx, disciplina in enumerate(disciplinas_sistema):
            # Abreviar nome da disciplina (pegar primeiras 4 letras em maiúsculo)
            abrev = disciplina[:4].upper() if len(disciplina) >= 4 else disciplina.upper()
            cell = ws.cell(row=1, column=col_disciplinas + idx)
            cell.value = abrev
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 5
        
        # TOTAL
        cell_total = ws.cell(row=1, column=col_disciplinas + len(disciplinas_sistema))
        cell_total.value = "TOTAL"
        cell_total.fill = total_fill
        cell_total.font = Font(bold=True, color="FFFFFF", size=11)
        cell_total.alignment = center_align
        cell_total.border = border
        ws.column_dimensions[cell_total.column_letter].width = 6
        
        # LINHA 2: Cabeçalhos fixos
        ws['A2'] = "TURMA"
        ws['A2'].fill = header_fill
        ws['A2'].font = header_font
        ws['A2'].alignment = center_align
        ws['A2'].border = border
        ws.column_dimensions['A'].width = 8
        
        ws['B2'] = "NOME"
        ws['B2'].fill = header_fill
        ws['B2'].font = header_font
        ws['B2'].alignment = center_align
        ws['B2'].border = border
        ws.column_dimensions['B'].width = 30
        
        # Agrupar dados por aluno
        alunos_unicos = df_historico[['matricula_aluno', 'nome_aluno', 'turma']].drop_duplicates()
        alunos_unicos = alunos_unicos.sort_values(['turma', 'nome_aluno'])
        
        linha_atual = 3
        
        for _, aluno in alunos_unicos.iterrows():
            matricula = aluno['matricula_aluno']
            nome = aluno['nome_aluno']
            turma = aluno['turma']
            
            # TURMA e NOME
            ws.cell(row=linha_atual, column=1).value = turma
            ws.cell(row=linha_atual, column=1).border = border
            ws.cell(row=linha_atual, column=1).alignment = center_align
            
            ws.cell(row=linha_atual, column=2).value = nome
            ws.cell(row=linha_atual, column=2).border = border
            
            # Faltas do aluno
            faltas_aluno = df_historico[df_historico['matricula_aluno'] == matricula]
            
            # Preencher faltas por dia
            for dia in range(1, dias_no_mes + 1):
                data_dia = f"{ano}-{mes:02d}-{dia:02d}"
                faltas_dia = faltas_aluno[faltas_aluno['data'] == data_dia]
                
                cell = ws.cell(row=linha_atual, column=col_offset + dia - 1)
                cell.value = len(faltas_dia) if not faltas_dia.empty else 0
                cell.alignment = center_align
                cell.border = border
            
            # Preencher totais por disciplina
            for idx, disciplina in enumerate(disciplinas_sistema):
                faltas_disc = faltas_aluno[faltas_aluno['disciplina'] == disciplina]
                cell = ws.cell(row=linha_atual, column=col_disciplinas + idx)
                cell.value = len(faltas_disc)
                cell.alignment = center_align
                cell.border = border
            
            # TOTAL GERAL
            total_faltas = len(faltas_aluno)
            cell_total = ws.cell(row=linha_atual, column=col_disciplinas + len(disciplinas_sistema))
            cell_total.value = total_faltas
            cell_total.alignment = center_align
            cell_total.border = border
            cell_total.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            linha_atual += 1
        
        # Salvar arquivo
        wb.save(caminho_arquivo)
        
        return True
        
    except Exception as e:
        raise Exception(f"Erro ao exportar para Excel: {str(e)}")


def obter_meses_com_faltas():
    """Retorna lista de meses/anos que possuem faltas registradas"""
    try:
        conn = sqlite3.connect(DB_HISTORICO)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT DISTINCT strftime('%Y-%m', data) as mes_ano
            FROM historico_faltas
            ORDER BY mes_ano DESC
        ''')
        
        resultados = cursor.fetchall()
        conn.close()
        
        meses = []
        for row in resultados:
            ano, mes = row[0].split('-')
            meses.append({'ano': int(ano), 'mes': int(mes), 'label': f"{mes}/{ano}"})
        
        return meses
        
    except Exception as e:
        return []
=======
        raise Exception(f"Erro ao buscar horário: {str(e)}")
>>>>>>> dff7806e9187c3a3a5bf23e75924bc777d34899c
